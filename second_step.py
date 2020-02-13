#获取每页内容
import time
from selenium import webdriver
import openpyxl

i = 2
browser = webdriver.Chrome()

def get_content(url):
    browser.get(url)
    time.sleep(3)

    result = browser.find_element_by_css_selector(".weibo-text").text
    post_time = browser.find_element_by_css_selector(".time").text
    
    out = ""

    for i in result:
            try:
                    print(i)
                    out = out + i
            except:
                    out = out + "[表情]"

    save(post_time,out)


def save(post_time,content):
   with open("result.txt","a") as f:
           f.write(post_time+"\n")
           try:
               f.write(content+"\n")
           except:
               f.write("【写入错误】\n")
           f.write("---------------------------------\n")

print("请先登录微博账户！并不要关闭此Chrome页面！")
time.sleep(40)

wb = openpyxl.load_workbook('link.xlsx')
ws = wb.active

for i in range(2,455):
    if ws.cell(row=i,column=1).value == "OK":
        save(ws.cell(row=i,column=3).value,ws.cell(row=i,column=2).value)
    else:
        get_content(ws.cell(row=i,column=1).value)
    print("完成第" + str(i-1) + "个\n")

quit()


