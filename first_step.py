#获取每条微博全文网址
import time
from selenium import webdriver
import openpyxl

browser = webdriver.Chrome()

print("请登录微博账户！并不注意不要关闭此Chrome页面！")
time.sleep(40)

#要爬取用户的手机版微博网址
usr_url = "https://m.weibo.cn/p/2304136418207965"
browser.get(usr_url)
print("已经加载完成")
time.sleep(30)


for i in range(1,10000):
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")

print("已经刷到底部")

wb = openpyxl.load_workbook('link.xlsx')
ws = wb.active

for i in range(0,10000):
    out = browser.find_elements_by_css_selector(".weibo-text")[i].text
    if out[len(out)-2:len(out)] == "全文":
        linkstr = browser.find_elements_by_xpath('//*[@class="weibo-text"]//a')[i].get_attribute("href")
        if linkstr.find('status') != -1:
            ws.cell(row=i+2,column=1,value = linkstr)
            wb.save('nicol.xlsx')
    else:
        ws.cell(row=i+2,column=1,value = "OK")
        real_out = ""
        for c in out:
            try:
                print(c)
                real_out = real_out + c
            except:
                real_out = real_out + "[表情]"
        ws.cell(row=i+2,column=2,value = real_out)
        ws.cell(row=i+2,column=3,value = browser.find_elements_by_css_selector(".time")[i].text)
        wb.save('link.xlsx')
    print("完成第"+ str(i+1) + "个\n")

quit()

